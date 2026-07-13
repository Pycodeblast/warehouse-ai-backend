import os
import shutil
from fastapi import UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.logger import logger
from app.models.product import Product


UPLOAD_FOLDER = "uploads"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

ALLOWED_EXTENSIONS = {
    ".pdf",
    ".xlsx",
    ".xls",
    ".csv",
    ".doc",
    ".docx",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".txt",
}


def get_file_type(extension: str):

    extension = extension.lower()

    if extension == ".pdf":
        return "PDF"

    elif extension in [".xls", ".xlsx"]:
        return "EXCEL"

    elif extension == ".csv":
        return "CSV"

    elif extension in [".doc", ".docx"]:
        return "WORD"

    elif extension in [".png", ".jpg", ".jpeg", ".webp"]:
        return "IMAGE"

    elif extension == ".txt":
        return "TEXT"

    return "OTHER"




def save_file(file: UploadFile):

    logger.info(f"Uploading file: {file.filename}")

    file_path = os.path.join(
        UPLOAD_FOLDER,
        file.filename
    )

    try:

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(
                file.file,
                buffer
            )

        logger.info(
            f"File uploaded successfully: {file.filename}"
        )

        return file_path


    except Exception as e:

        logger.error(
            f"File upload failed: {str(e)}"
        )

        raise



def import_products_from_excel(
    file_path: str,
    db: Session
):

    workbook = load_workbook(file_path)
    sheet = workbook.active

    imported_count = 0
    skipped_count = 0
    duplicate_skus = []

    # Skip header row
    for row in sheet.iter_rows(min_row=2, values_only=True):

        name, sku, quantity, price = row[:4]

        # Skip empty rows
        if not sku:
            continue

        # Check whether SKU already exists
        existing_product = (
            db.query(Product)
            .filter(Product.sku == sku)
            .first()
        )

        if existing_product:
            skipped_count += 1
            duplicate_skus.append(sku)
            continue

        product = Product(
            name=name,
            sku=sku,
            quantity=quantity,
            price=price
        )

        db.add(product)
        imported_count += 1

    db.commit()

    logger.info(
        f"Imported: {imported_count}, Skipped: {skipped_count}"
    )

    return {
        "imported": imported_count,
        "skipped": skipped_count,
        "duplicate_skus": duplicate_skus
    }